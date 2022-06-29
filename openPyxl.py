from openpyxl import load_workbook, Workbook
from openpyxl.styles import Color, Font, PatternFill
import os
import socket
import time
from datetime import datetime

class listCalledExcel():
    def __init__(self, filename):
        self.filename = filename
        self.wb = self.createPlan(self.filename)
        self.ws = self.wb.active
        self.local = filename

    # RETORNA NOME E CAMINHO DA PLANILHA
    def getFilename(self):
        return str

    # RETORNA PLANILHA FORMATADA
    def getFormatedDataPlan(self):
        calledList = self.getListCalled()
        for x in range(len(calledList)):
            print("\nChamado", x + 1)
            for y in calledList[x]:
                if calledList[x][y] == "None":
                    pass
                else:
                    print(y, ":", calledList[x][y])

    def getFormatedDataPlanHisotry(self):
        calledList = self.getListHistory()
        for x in range(len(calledList)):
            print("\nChamado", x + 1)
            for y in calledList[x]:
                print(y, ":", calledList[x][y])

    # RETORNA DICIONARIOS PRONTOS PARA ABERTURA DE CHAMADO.
    # RETORNA APENAS LINHAS SEM CDCHAMADO PREENCHIDO
    def getListCalled(self):
        listDataCalled = []
        for listRow in range(3, self.ws.max_row, 1):
            dictRow = {}
            for listColum in range(self.ws.max_column):
                if self.ws.cell(row=listRow+1, column=1).value == None \
                        or self.ws.cell(row=listRow+1, column=3).value == None \
                        or self.ws.cell(row=listRow + 1, column=4).value != None\
                        or self.ws.cell(row=listRow+1, column=5).value == None\
                        or self.ws.cell(row=listRow+1, column=6).value == None:
                    break
                elif self.ws.cell(row=listRow+1, column=4).value != None:
                    break
                else:
                    dictRow.update({str(self.ws.cell(row=3, column=listColum+1).value): str(self.ws.cell(row=listRow+1, column=listColum+1).value)})
            if len(dictRow) != 0:
                listDataCalled.append(dictRow)
            else:
                pass
        return listDataCalled

    # RETORNA DICIONARIOS COM CHAMADOS JA ABERTOS
    # RETORNA LINHAS COM E SEM CDCHAMADO PREENCHIDO
    def getListHistory(self):
        listDataCalled = []
        for listRow in range(3, self.ws.max_row, 1):
            dictRow = {}
            for listColum in range(self.ws.max_column):
                if self.ws.cell(row=listRow+1, column=1).value == None \
                        or self.ws.cell(row=listRow+1, column=3).value == None \
                        or self.ws.cell(row=listRow + 1, column=4).value == None\
                        or self.ws.cell(row=listRow+1, column=5).value == None\
                        or self.ws.cell(row=listRow+1, column=6).value == None:
                    break
                elif self.ws.cell(row=listRow+1, column=4).value == None:
                    pass
                else:
                    dictRow.update({str(self.ws.cell(row=3, column=listColum+1).value): str(self.ws.cell(row=listRow+1, column=listColum+1).value)})
            if len(dictRow) != 0:
                listDataCalled.append(dictRow)
            else:
                pass
        return listDataCalled

    # CADASTRAR NUMEROS DE CHAMADOS NA PLANILHA DE ABERTURA
    def sendNumCalled(self, listNumCalled):
        def burn(wb, ws):
            for x in range(len(listNumCalled)):
                if listNumCalled[x] != 'None':
                    ws['D' + str(x + 4)] = listNumCalled[x]
                else:
                    pass
            wb.save(f'\\\\'+ socket.gethostname() + '\\C$\\Users\\' + os.getlogin() + '\\Documents\\GhostNumbers\\' + self.filename)

        # CRIA ARQUIVO DE LOG COM NUMEROS DE CHAMADO ABERTOS
        def createFileCalled(listCalled):
            NoteFile = open(f'\\\\'+ socket.gethostname() + '\\C$\\Users\\' + os.getlogin() + '\\Documents\\GhostNumbers\\logs\\' + str(datetime.now().strftime('%Y.%m.%d %H-%M-%S')) + '.txt', "a")
            for x in range(len(listCalled)):
                NoteFile.write(listCalled[x] + '\n')
            NoteFile.close()
        #createFileCalled(listNumCalled)

        while True:
            try:
                wb = self.createPlan(self.filename)
                ws = wb.active
                wb.save(f'\\\\'+ socket.gethostname() + '\\C$\\Users\\' + os.getlogin() + '\\Documents\\GhostNumbers\\' + self.filename)
                burn(wb, ws)
                break
            except PermissionError:
                input("\n FECHAR PLANILHA ABERTA \n")
                time.sleep(5)


    # VERIFICA E CRIA PASTA E ARQUIVO
    # SE NAO EXISTIR PASTA OU ARQUIVO SERÃO CRIADOS.
    def createPlan(self, filename):
        try:
            #wb = load_workbook('C:\\Users\\' + os.getlogin() + '\\Documents\\GhostNumbers\\'+filename)
            wb = load_workbook(f'\\\\'+ socket.gethostname() + '\\C$\\Users\\' + os.getlogin() + '\\Documents\\GhostNumbers\\' + filename)
            ws = wb.active
            return wb
        except FileNotFoundError:
            try:
                os.mkdir(f'\\\\'+ socket.gethostname() + '\\C$\\Users\\' + os.getlogin() + '\\Documents\\GhostNumbers\\')
                os.mkdir(f'\\\\'+ socket.gethostname() + '\\C$\\Users\\' + os.getlogin() + '\\Documents\\GhostNumbers\\logs')
            except FileExistsError:
                pass
            wb = Workbook()
            ws = wb.active
            ws.title = 'PlanilhaAberturaDeChamados'
            fontAll = Font(color="FFFFFF", bold=True)
            fontCampTempl = Font(color="FFFFFF", bold=True)
            fillAll = PatternFill('solid', start_color="00041a")
            fillCampTempl = PatternFill('solid', start_color="0022af")
            ws['A1'] = 'cdresponsavel'
            ws['A1'].fill = fillAll
            ws['A1'].font = fontAll
            ws['A2'] = 'CONTATO'
            ws['A2'].fill = fillAll
            ws['A2'].font = fontAll
            ws['A3'] = 'cdcontato'
            ws['A3'].fill = fillAll
            ws['A3'].font = fontAll
            ws['B1'] = 'Cod. Responsavel. EX.: 348'
            ws['B1'].fill = fillCampTempl
            ws['B1'].font = fontCampTempl
            ws['B2'] = 'INTERNO / INCIDENTE'
            ws['B2'].fill = fillAll
            ws['B2'].font = fontAll
            ws['B3'] = 'cdcategoria'
            ws['B3'].fill = fillAll
            ws['B3'].font = fontAll
            ws['C1'] = 'cdequipe'
            ws['C1'].fill = fillAll
            ws['C1'].font = fontAll
            ws['C2'] = 'DESCRIÇÃO DE ABERTURA'
            ws['C2'].fill = fillAll
            ws['C2'].font = fontAll
            ws['C3'] = 'dschamado'
            ws['C3'].fill = fillAll
            ws['C3'].font = fontAll
            ws['D1'] = 'Cod. equipe. EX.: 20'
            ws['D1'].fill = fillCampTempl
            ws['D1'].font = fontCampTempl
            ws['D2'] = 'NUMERO CHAMADO'
            ws['D2'].fill = fillAll
            ws['D2'].font = fontAll
            ws['D3'] = 'cdchamado'
            ws['D3'].fill = fillAll
            ws['D3'].font = fontAll
            ws['E2'] = 'DESCRIÇÃO DE ACOMPANHAMENTO'
            ws['E2'].fill = fillAll
            ws['E2'].font = fontAll
            ws['E3'] = 'dsacompanhamento'
            ws['E3'].fill = fillAll
            ws['E3'].font = fontAll
            ws['F2'] = 'DURAÇÃO DE ATENDIMENTO 000:00'
            ws['F2'].fill = fillAll
            ws['F2'].font = fontAll
            ws['F3'] = 'hrduracao'
            ws['F3'].fill = fillAll
            ws['F3'].font = fontAll
            local = f'\\\\'+ socket.gethostname() + '\\C$\\Users\\' + os.getlogin() + '\\Documents\\GhostNumbers\\' + filename
            wb.save(local)
            print(local)
            return wb